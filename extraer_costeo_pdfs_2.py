# -*- coding: utf-8 -*-
"""
IDENTIFICADOR Y EXTRACTOR DE COSTEO UNITARIO DESDE PDFs — SECOP-II / Claude Sonnet
====================================================================
Lee TODOS los PDFs de la carpeta, envía cada uno a Claude Sonnet 4.5
y extrae: municipio/subregión, precio, cantidad, costo unitario,
descripción del contrato, año y código del contrato.

LIBRERIAS O PAQUETES:
    pip install anthropic pandas openpyxl
    python extraer_costeo_pdfs.py

    # Para probar con pocos archivos primero:
    python extraer_costeo_pdfs.py --limite 5

SALIDA:
    costeo_unitario_contratos.xlsx
    extraccion_costeo.log
"""

import os
import re
import sys
import json
import time
import base64
import logging
import argparse
from pathlib import Path
from datetime import datetime
from typing import Optional

import pandas as pd
import anthropic
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# _____________________________________________________________________________

# CONFIGURACIÓN  — EDITA ESTAS VARIABLES

# Ruta donde están los PDFs (todos sueltos en esta carpeta)
RUTA_PDFS = r"C:\Users\velez\Documents\Consultorias_2026\ART_Costeo\Datos\Nueva_Estrategia\Fase_II\Pdf_mcp_final"

# Usando API key de Anthropic
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

# Modelo Claude a usar
MODELO_CLAUDE = "claude-sonnet-4-5"

# Archivos de salida (se crean en el directorio desde donde ejecutas el script)
EXCEL_SALIDA = "costeo_unitario_contratos.xlsx"
LOG_FILE     = "extraccion_costeo.log"

# Parámetros de procesamiento
MAX_TAMANO_MB        = 30     # Claude acepta hasta 30 MB por PDF
MAX_PAGINAS_PDF      = 100    # Claude acepta hasta 100 páginas
PAUSA_ENTRE_LLAMADAS = 70     # segundos entre llamadas (Tier 1 requiere ~70s por PDF grande)
MAX_REINTENTOS       = 5      # reintentos ante fallo
PAUSA_RATE_LIMIT     = 90     # segundos a esperar al detectar 429
CHECKPOINT_CADA      = 5      # guarda Excel parcial cada N PDFs
MAX_TOKENS_RESPUESTA = 1500


# =============================================================================
# LOGGING
# =============================================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# =============================================================================
# PROMPT PARA CLAUDE
# =============================================================================

PROMPT = """Eres un analista experto en contratación pública colombiana (SECOP-II). \
Tu tarea es leer el documento contractual adjunto (Estudio Previo, Anexo Técnico o \
similar) y extraer la información clave para calcular el COSTEO UNITARIO del contrato.

Extrae los siguientes campos. Si un dato NO aparece explícitamente en el documento, \
pon `null` (NO inventes valores).

1. **codigo_contrato**: identificador del proceso SECOP-II (ej: CO1.NTC.1958826), \
número interno del contrato (ej: CCFV-065-2024), o número de proceso de selección \
(ej: LP-001-2024). Si aparecen varios, concatenalos con " | ". null si no encuentras ninguno.

2. **descripcion_contrato**: objeto del contrato tal como aparece en el documento \
(texto completo del objeto, máximo 500 caracteres).

3. **municipio**: municipio(s) o lugar(es) específicos de ejecución del contrato. \
Si son varios, sepáralos con "; ". Usa los nombres exactos como aparecen en el documento.

4. **departamento**: departamento(s) correspondiente(s) (Colombia). Separa con "; " si son varios.

5. **subregion_impacto**: si el documento menciona explícitamente una subregión PDET \
(ej: "Alto Patía - Norte del Cauca", "Catatumbo", "Sur de Bolívar", "Macarena - Guaviare") \
o una región de impacto específica, indícala. null si no aparece.

6. **year_contrato**: año del contrato (formato YYYY). Si aparece fecha completa, extrae \
el año. Si hay varias fechas (firma, inicio, terminación), usa la de FIRMA o SUSCRIPCIÓN. null si no aparece.

7. **precio_cop**: valor total del contrato en pesos colombianos (COP), como número entero \
sin puntos, comas ni símbolo $. Ejemplo: 7035424595. Si el valor está en otra moneda, \
extráelo tal cual y anota la moneda en `moneda`.

8. **moneda**: "COP", "USD", "EUR", etc. Por defecto "COP".

9. **cantidad**: cantidad física que el contrato entrega/ejecuta, en su unidad natural. \
Por ejemplo: 683600 (hectáreas), 12.5 (km de vía), 240 (familias beneficiarias), \
1 (plan formulado), 150 (cupos de formación). Usa número decimal con punto. null si no aparece.

10. **unidad_cantidad**: unidad física de la cantidad (ej: "hectárea", "km", "familia", \
"plan", "organización apoyada", "cupo de formación", "estudiante", "kit entregado"). \
null si no aparece.

11. **costo_unitario_cop**: costo por unidad en COP. Si el documento lo reporta explícitamente, \
úsalo. Si no, calcúlalo como `precio_cop / cantidad`. Redondea a entero. null si no se puede \
calcular (faltan datos).

12. **confianza**: tu confianza en la extracción — "alta", "media" o "baja". \
"baja" si el documento es ambiguo, está escaneado con mala resolución, o no encontraste \
la mayoría de los campos clave.

13. **observaciones**: nota del analista (máximo 300 caracteres). Por ejemplo: \
"Contrato incluye interventoría dentro del valor"; "Cantidad aproximada"; \
"El documento es una adición, no el contrato original"; "No especifica cantidad, solo valor total".

FORMATO DE RESPUESTA
────────────────────
Responde ÚNICAMENTE con un objeto JSON válido. Sin texto antes ni después, sin backticks, \
sin explicaciones:

{
  "codigo_contrato": "...",
  "descripcion_contrato": "...",
  "municipio": "...",
  "departamento": "...",
  "subregion_impacto": "...",
  "year_contrato": 2024,
  "precio_cop": 0,
  "moneda": "COP",
  "cantidad": 0.0,
  "unidad_cantidad": "...",
  "costo_unitario_cop": 0,
  "confianza": "alta",
  "observaciones": "..."
}
"""


# =============================================================================
# FUNCIONES
# =============================================================================

def pdf_apto(path_pdf: Path) -> tuple[bool, str]:
    """Verifica que un PDF cumpla los límites de la API de Anthropic."""
    if not path_pdf.exists():
        return False, "archivo no existe"
    size_mb = path_pdf.stat().st_size / (1024 * 1024)
    if size_mb > MAX_TAMANO_MB:
        return False, f"demasiado grande ({size_mb:.1f} MB > {MAX_TAMANO_MB} MB)"
    if size_mb < 0.001:
        return False, "archivo vacío"
    return True, f"{size_mb:.2f} MB"


def llamar_claude_pdf(client: anthropic.Anthropic, path_pdf: Path) -> Optional[dict]:
    """
    Envía el PDF a Claude y devuelve el JSON parseado, o None si falla.
    Usa la API nativa de Messages con input tipo 'document' (base64).
    """
    with open(path_pdf, "rb") as f:
        pdf_b64 = base64.standard_b64encode(f.read()).decode("utf-8")

    contenido = [
        {
            "type": "document",
            "source": {
                "type": "base64",
                "media_type": "application/pdf",
                "data": pdf_b64,
            },
        },
        {"type": "text", "text": PROMPT},
    ]

    for intento in range(1, MAX_REINTENTOS + 1):
        try:
            msg = client.messages.create(
                model=MODELO_CLAUDE,
                max_tokens=MAX_TOKENS_RESPUESTA,
                messages=[{"role": "user", "content": contenido}],
            )
            raw = msg.content[0].text.strip()
            raw = raw.replace("```json", "").replace("```", "").strip()

            # Si el modelo añade texto alrededor, extraer el bloque JSON
            m = re.search(r"\{.*\}", raw, re.DOTALL)
            if m:
                raw = m.group(0)

            resultado = json.loads(raw)

            if hasattr(msg, "usage"):
                log.info(
                    f"    tokens in={msg.usage.input_tokens} "
                    f"out={msg.usage.output_tokens}"
                )
            return resultado

        except json.JSONDecodeError as e:
            log.warning(f"  JSON inválido (intento {intento}): {e}")
            time.sleep(2 * intento)

        except anthropic.RateLimitError:
            log.warning(f"  Rate limit (intento {intento}), esperando {PAUSA_RATE_LIMIT}s...")
            time.sleep(PAUSA_RATE_LIMIT)

        except anthropic.APIStatusError as e:
            log.warning(f"  Error API {e.status_code} (intento {intento}): {e.message}")
            if e.status_code == 400:
                # PDF corrupto, encriptado o demasiado largo — no reintentar
                return None
            time.sleep(3 * intento)

        except Exception as e:
            log.warning(f"  Error inesperado (intento {intento}): {type(e).__name__}: {e}")
            time.sleep(3 * intento)

    log.error(f"  Falla tras {MAX_REINTENTOS} intentos: {path_pdf.name}")
    return None


def formatear_excel(path_excel: Path):
    """Aplica formato visual al Excel resultante."""
    wb = load_workbook(path_excel)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_header
        cell.border = border

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

    # Anchos de columna
    anchos = {
        "archivo_pdf": 35,
        "codigo_contrato": 24,
        "descripcion_contrato": 55,
        "municipio": 28,
        "departamento": 18,
        "subregion_impacto": 24,
        "year_contrato": 10,
        "precio_cop": 16,
        "moneda": 9,
        "cantidad": 12,
        "unidad_cantidad": 18,
        "costo_unitario_cop": 16,
        "confianza": 11,
        "observaciones": 40,
        "fecha_extraccion": 18,
    }
    for i, col in enumerate(ws[1], start=1):
        letra = get_column_letter(i)
        ws.column_dimensions[letra].width = anchos.get(col.value, 15)

    # Formato de número para columnas monetarias
    cols_moneda = ["precio_cop", "costo_unitario_cop"]
    for i, col in enumerate(ws[1], start=1):
        if col.value in cols_moneda:
            letra = get_column_letter(i)
            for cell in ws[letra][1:]:
                if cell.value is not None:
                    cell.number_format = '"$"#,##0'

    # Colorear filas por confianza
    fills_conf = {
        "alta":  PatternFill("solid", fgColor="E8F5E9"),
        "media": PatternFill("solid", fgColor="FFF9C4"),
        "baja":  PatternFill("solid", fgColor="FFEBEE"),
    }
    col_conf_idx = None
    for i, col in enumerate(ws[1], start=1):
        if col.value == "confianza":
            col_conf_idx = i
            break

    if col_conf_idx:
        for row in ws.iter_rows(min_row=2):
            val = row[col_conf_idx - 1].value
            fill = fills_conf.get(str(val).lower() if val else None)
            if fill:
                for cell in row:
                    cell.fill = fill

    wb.save(path_excel)


def cargar_resultados_previos(path: Path) -> tuple[list[dict], set[str]]:
    """
    Si existe el Excel de salida, lo carga y devuelve:
      - la lista de resultados ya procesados
      - un set con los nombres de archivo ya completados
    Así el script puede reanudar donde se quedó.
    """
    if not path.exists():
        return [], set()

    try:
        df_prev = pd.read_excel(path, sheet_name="Costeo_Unitario")
        if df_prev.empty or "archivo_pdf" not in df_prev.columns:
            return [], set()

        # Convertir NaN a None para compatibilidad con JSON/dict
        df_prev = df_prev.where(pd.notna(df_prev), None)
        resultados_previos = df_prev.to_dict(orient="records")
        procesados = set(df_prev["archivo_pdf"].dropna().astype(str).tolist())
        return resultados_previos, procesados

    except Exception as e:
        log.warning(f"No se pudo leer el Excel previo ({path.name}): {e}")
        log.warning("Se empezará desde cero.")
        return [], set()


def guardar_excel(resultados: list[dict], path: Path):
    """Guarda la lista de resultados en Excel con formato."""
    # Orden preferido de columnas
    orden = [
        "archivo_pdf",
        "codigo_contrato",
        "descripcion_contrato",
        "municipio",
        "departamento",
        "subregion_impacto",
        "year_contrato",
        "precio_cop",
        "moneda",
        "cantidad",
        "unidad_cantidad",
        "costo_unitario_cop",
        "confianza",
        "observaciones",
        "fecha_extraccion",
    ]
    df = pd.DataFrame(resultados)
    # Añadir columnas que falten para mantener orden
    for c in orden:
        if c not in df.columns:
            df[c] = None
    df = df[orden]
    df.to_excel(path, index=False, sheet_name="Costeo_Unitario")
    try:
        formatear_excel(path)
    except Exception as e:
        log.warning(f"No se pudo aplicar formato al Excel: {e}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Extrae costeo unitario de PDFs SECOP con Claude."
    )
    parser.add_argument("--ruta-pdfs", default=RUTA_PDFS,
                        help="Carpeta con los PDFs")
    parser.add_argument("--salida", default=EXCEL_SALIDA,
                        help="Nombre del Excel de salida")
    parser.add_argument("--limite", type=int, default=None,
                        help="Procesar solo los primeros N archivos (prueba)")
    parser.add_argument("--rehacer", action="store_true",
                        help="Ignora el Excel existente y reprocesa TODOS los PDFs desde cero")
    args = parser.parse_args()

    # Validaciones
    if not ANTHROPIC_API_KEY:
        log.error("No se encontró la API key. Define la variable de entorno "
                  "ANTHROPIC_API_KEY o edita el código.")
        log.error('  En PowerShell: $env:ANTHROPIC_API_KEY = "sk-ant-..."')
        sys.exit(1)

    ruta_raiz = Path(args.ruta_pdfs)
    if not ruta_raiz.exists():
        log.error(f"No existe la ruta de PDFs: {ruta_raiz}")
        sys.exit(1)

    # Listar PDFs (recursivo — funciona aunque haya subcarpetas o no)
    pdfs = sorted(ruta_raiz.rglob("*.pdf"))
    # Eliminar duplicados por nombre de ruta (por si acaso)
    pdfs = list(dict.fromkeys(pdfs))

    if not pdfs:
        log.error(f"No se encontraron PDFs en {ruta_raiz}")
        sys.exit(1)

    # --- REANUDACIÓN: cargar Excel previo si existe -------------------------
    excel_path = Path(args.salida)
    resultados, procesados = cargar_resultados_previos(excel_path)

    # Filtrar PDFs ya procesados (comparando por nombre de archivo)
    if procesados and not args.rehacer:
        pdfs_pendientes = [p for p in pdfs if p.name not in procesados]
        ya_procesados = len(pdfs) - len(pdfs_pendientes)
    else:
        pdfs_pendientes = pdfs
        ya_procesados = 0
        if args.rehacer:
            resultados = []
            procesados = set()

    log.info("=" * 70)
    log.info("EXTRACCIÓN DE COSTEO UNITARIO CON CLAUDE SONNET")
    log.info("=" * 70)
    log.info(f"  Ruta de PDFs:    {ruta_raiz}")
    log.info(f"  PDFs detectados: {len(pdfs)}")
    if ya_procesados > 0:
        log.info(f"  Ya procesados:   {ya_procesados} (se omiten)")
        log.info(f"  Pendientes:      {len(pdfs_pendientes)}")
    log.info(f"  Modelo:          {MODELO_CLAUDE}")
    log.info(f"  Excel salida:    {args.salida}")
    log.info("=" * 70)

    if not pdfs_pendientes:
        log.info("Todos los PDFs ya están procesados. Nada por hacer.")
        log.info(f"Si quieres reprocesar todo, ejecuta con --rehacer")
        return

    pdfs = pdfs_pendientes

    if args.limite:
        pdfs = pdfs[: args.limite]
        log.info(f"Modo prueba: se procesarán solo {len(pdfs)} PDFs pendientes")

    # Cliente Anthropic
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    total = len(pdfs)
    exitos = 0
    fallos = 0

    for idx, pdf in enumerate(pdfs, start=1):
        log.info(f"\n[{idx}/{total}] {pdf.name}")

        # Validar tamaño/integridad
        ok, info = pdf_apto(pdf)
        if not ok:
            log.warning(f"  Descartado: {info}")
            resultados.append({
                "archivo_pdf":        pdf.name,
                "confianza":          "baja",
                "observaciones":      f"PDF descartado: {info}",
                "fecha_extraccion":   datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            })
            fallos += 1
            continue

        log.info(f"  Enviando a Claude ({info})...")
        extraccion = llamar_claude_pdf(client, pdf)

        if extraccion is None:
            log.warning(f"  Sin respuesta del API — NO se guarda (se reintentará en próxima ejecución)")
            fallos += 1
            # NO se añade a resultados: así, al reanudar, se volverá a intentar
        else:
            # Armar fila completa
            fila = {
                "archivo_pdf":          pdf.name,
                "codigo_contrato":      extraccion.get("codigo_contrato"),
                "descripcion_contrato": extraccion.get("descripcion_contrato"),
                "municipio":            extraccion.get("municipio"),
                "departamento":         extraccion.get("departamento"),
                "subregion_impacto":    extraccion.get("subregion_impacto"),
                "year_contrato":        extraccion.get("year_contrato"),
                "precio_cop":           extraccion.get("precio_cop"),
                "moneda":               extraccion.get("moneda"),
                "cantidad":             extraccion.get("cantidad"),
                "unidad_cantidad":      extraccion.get("unidad_cantidad"),
                "costo_unitario_cop":   extraccion.get("costo_unitario_cop"),
                "confianza":            extraccion.get("confianza"),
                "observaciones":        extraccion.get("observaciones"),
                "fecha_extraccion":     datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }

            # Si el modelo no calcula el costo unitario pero hay precio y cantidad, entonces se calcula
            try:
                if (not fila["costo_unitario_cop"]
                        and fila["precio_cop"]
                        and fila["cantidad"]
                        and float(fila["cantidad"]) > 0):
                    fila["costo_unitario_cop"] = round(
                        float(fila["precio_cop"]) / float(fila["cantidad"])
                    )
            except (TypeError, ValueError):
                pass

            resultados.append(fila)
            exitos += 1
            log.info(
                f"  OK {fila.get('municipio') or '-'} | "
                f"${fila.get('precio_cop') or '-'} | "
                f"{fila.get('cantidad') or '-'} {fila.get('unidad_cantidad') or ''} | "
                f"conf={fila.get('confianza')}"
            )

        # Pausa entre llamadas
        time.sleep(PAUSA_ENTRE_LLAMADAS)

        # Checkpoint
        if idx % CHECKPOINT_CADA == 0 or idx == total:
            guardar_excel(resultados, Path(args.salida))
            log.info(f"  [Checkpoint] {idx}/{total} guardados en {args.salida}")

    # Guardado final
    guardar_excel(resultados, Path(args.salida))

    # Resumen
    log.info("\n" + "=" * 70)
    log.info("RESUMEN FINAL")
    log.info("=" * 70)
    log.info(f"  PDFs procesados:       {total}")
    log.info(f"  Extracciones exitosas: {exitos}")
    log.info(f"  Fallos:                {fallos}")

    df = pd.DataFrame(resultados)
    if not df.empty and "costo_unitario_cop" in df.columns:
        con_costo = df["costo_unitario_cop"].notna().sum()
        log.info(f"  Con costo unitario:    {con_costo}")
        if "confianza" in df.columns:
            for nivel in ["alta", "media", "baja"]:
                n = (df["confianza"] == nivel).sum()
                log.info(f"  Confianza '{nivel}': {n}")

    log.info(f"\n  Archivo final: {Path(args.salida).resolve()}")
    log.info("=" * 70)


if __name__ == "__main__":
    main()
