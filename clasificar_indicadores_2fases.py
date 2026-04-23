"""
Clasificador en 2 Fases: Indicadores, Segmento, Familia, Producto (SECOP II)
================================================================================
Estrategia de mínimo consumo de tokens:

  FASE 1  Clasifica cada indicador a la variable SEGMENTO más cercano.
          Catálogo: 57 segmentos únicos, estableciendo un prompt pequeño y fijo.
          Batch: 20 indicadores por llamada → 31 llamadas.

  FASE 2  Dentro del segmento asignado, clasifica al PRODUCTO más cercano.
          Indicadores que están en el mismo segmento se agrupan en un único lote.

Archivos de entrada (misma carpeta que el script):
  - data_priorizados_Indicadores.xlsx   (hoja: Indicador_asociado, col B)
  - jerarquia_completa.xlsx

Archivo generado:
  - indicadores_clasificados_2fases.xlsx
"""

import os
import anthropic
import pandas as pd
import json
import time
from pathlib import Path
from collections import defaultdict

# CONFIGURACIÓN

ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")  # set in env: export ANTHROPIC_API_KEY="sk-ant-..."

ARCHIVO_INDICADORES = "data_priorizados_Indicadores.xlsx"
HOJA_INDICADORES    = "Indicador_asociado"
ARCHIVO_JERARQUIA   = "jerarquia_completa.xlsx"
ARCHIVO_SALIDA      = "indicadores_clasificados_2fases.xlsx"

BATCH_SIZE_F1       = 20   # indicadores por llamada en Fase 1
BATCH_SIZE_F2       = 15   # indicadores por llamada en Fase 2 (catálogos más largos)
PAUSA               = 0.4  # segundos entre llamadas
MAX_REINTENTOS      = 3
# ─────────────────────────────────────────────

#  UTILIDADES


def llamar_api(client, prompt, max_tokens=1500):
    """Llamada a la API con reintentos."""
    for intento in range(1, MAX_REINTENTOS + 1):
        try:
            msg = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=max_tokens,
                messages=[{"role": "user", "content": prompt}]
            )
            raw = msg.content[0].text.replace("```json", "").replace("```", "").strip()
            return json.loads(raw)

        except json.JSONDecodeError as e:
            print(f"\n  ⚠️  JSON inválido (intento {intento}): {e}")
            time.sleep(2 * intento)
        except anthropic.RateLimitError:
            print(f"\n  ⚠️  Rate limit (intento {intento}), esperando 30s...")
            time.sleep(30)
        except Exception as e:
            print(f"\n  ⚠️  Error (intento {intento}): {e}")
            time.sleep(3 * intento)

    return None   # señal de fallo tras todos los reintentos

# ══════════════════════════════════════════════

#  CARGA DE DATOS


def cargar_datos():
    # ── Indicadores (columna B, hoja Indicador_asociado) ──
    df_ind = pd.read_excel(ARCHIVO_INDICADORES, sheet_name=HOJA_INDICADORES)
    col_b  = df_ind.columns[1]
    indicadores = (
        df_ind[col_b]
        .astype(str)
        .str.replace(r"\n", " ", regex=True)
        .str.strip()
        .tolist()
    )
    indicadores = [i for i in indicadores if i and i.lower() not in ("nan", col_b.lower())]

    # ── Jerarquía completa ──
    df_jer = (
        pd.read_excel(ARCHIVO_JERARQUIA)
        [["nombre_segmento", "nombre_familia", "nombre_clase", "nombre_producto"]]
        .dropna()
        .drop_duplicates()
    )

    print(f"📋 Indicadores cargados  : {len(indicadores)}")
    print(f"🗂  Segmentos únicos      : {df_jer['nombre_segmento'].nunique()}")
    print(f"🗂  Familias únicas       : {df_jer['nombre_familia'].nunique()}")
    print(f"📦 Productos únicos      : {df_jer['nombre_producto'].nunique()}")

    return indicadores, df_jer


# ══════════════════════════════════════════════
#  FASE 1 — Indicador → Segmento


def prompt_fase1(batch, segmentos_txt):
    batch_txt = "\n".join(f"{i+1}. {ind}" for i, ind in enumerate(batch))
    return f"""Eres experto en clasificación de bienes y servicios del sector público colombiano \
usando el estándar UNSPSC adaptado por Colombia Compra Eficiente para SECOP II.

SEGMENTOS DISPONIBLES:
{segmentos_txt}

Asocia cada indicador de política pública PDET con el SEGMENTO semánticamente más cercano.
Reglas:
1. Usa EXACTAMENTE el texto del segmento tal como aparece en la lista.
2. Si ninguno es razonablemente cercano, escribe "Sin clasificar".
3. Responde ÚNICAMENTE con un JSON array válido, sin texto adicional ni backticks.

INDICADORES:
{batch_txt}

Formato (mismo orden que los indicadores):
[{{"indicador":"texto exacto","segmento":"segmento exacto"}}]"""


def fase1(client, indicadores, df_jer):
    segmentos     = sorted(df_jer["nombre_segmento"].unique().tolist())
    segmentos_txt = "\n".join(segmentos)

    batches = [indicadores[i:i+BATCH_SIZE_F1]
               for i in range(0, len(indicadores), BATCH_SIZE_F1)]
    total   = len(batches)
    results = []

    print(f"\n{'='*55}")
    print(f"  FASE 1: Indicador → Segmento  ({total} lotes de ≤{BATCH_SIZE_F1})")
    print(f"{'='*55}")

    for i, batch in enumerate(batches):
        print(f"  ⏳ Lote {i+1:02d}/{total} ({len(batch)} indicadores)...", end=" ", flush=True)
        prompt = prompt_fase1(batch, segmentos_txt)
        data   = llamar_api(client, prompt, max_tokens=max(1600, len(batch) * 80))

        if data:
            results.extend(data)
            print(f"✅ {len(data)} ok")
        else:
            for ind in batch:
                results.append({"indicador": ind, "segmento": "Sin clasificar"})
            print("❌ fallido → 'Sin clasificar'")

        time.sleep(PAUSA)

    df_f1 = pd.DataFrame(results)[["indicador", "segmento"]]

    sc = (df_f1["segmento"] == "Sin clasificar").sum()
    print(f"\n  Fase 1 completa — clasificados: {len(df_f1)-sc}/{len(df_f1)}")
    return df_f1


# ══════════════════════════════════════════════
#  FASE 2 — Indicador → Producto (dentro del segmento)


def prompt_fase2(batch, productos_txt, segmento):
    batch_txt = "\n".join(f"{i+1}. {ind}" for i, ind in enumerate(batch))
    return f"""Eres experto en clasificación de bienes y servicios del sector público colombiano \
usando el estándar UNSPSC adaptado por Colombia Compra Eficiente para SECOP II.

SEGMENTO: {segmento}

PRODUCTOS DISPONIBLES EN ESTE SEGMENTO:
{productos_txt}

Asocia cada indicador con el PRODUCTO más específico y semánticamente más cercano.
Reglas:
1. Usa excactamente el texto del producto tal como aparece en la lista.
2. Si ninguno es razonablemente cercano, escribe "Sin clasificar".
3. Responde unicamente con un JSON array válido, sin texto adicional ni backticks.

INDICADORES:
{batch_txt}

Formato (mismo orden que los indicadores):
[{{"indicador":"texto exacto","producto":"producto exacto"}}]"""


def fase2(client, df_f1, df_jer):
    """
    Agrupa los indicadores por segmento asignado en Fase 1.
    Por cada segmento, construye UN catálogo (solo productos de ese segmento)
    y lanza lotes de hasta BATCH_SIZE_F2 indicadores.
    Esto minimiza tokens: el catálogo se construye una vez por segmento.
    """
    # Índice: segmento → lista de productos (texto plano)
    prod_por_seg = (
        df_jer
        .groupby("nombre_segmento")["nombre_producto"]
        .apply(lambda s: sorted(s.dropna().unique().tolist()))
        .to_dict()
    )

    # Agrupar indicadores por segmento
    grupos = defaultdict(list)
    for _, row in df_f1.iterrows():
        grupos[row["segmento"]].append(row["indicador"])

    total_segmentos = len(grupos)
    results = []
    seg_idx = 0

    print(f"\n{'='*55}")
    print(f"  FASE 2: Indicador → Producto  ({total_segmentos} segmentos distintos)")
    print(f"{'='*55}")

    for segmento, indicadores_seg in sorted(grupos.items()):
        seg_idx += 1

        if segmento == "Sin clasificar":
            for ind in indicadores_seg:
                results.append({"indicador": ind, "producto": "Sin clasificar"})
            print(f"  [{seg_idx:02d}/{total_segmentos}] Sin clasificar ({len(indicadores_seg)} ind.) → omitido")
            continue

        productos   = prod_por_seg.get(segmento, [])
        if not productos:
            for ind in indicadores_seg:
                results.append({"indicador": ind, "producto": "Sin clasificar"})
            print(f"  [{seg_idx:02d}/{total_segmentos}] {segmento[:50]} — sin productos → omitido")
            continue

        productos_txt = "\n".join(productos)
        batches = [indicadores_seg[i:i+BATCH_SIZE_F2]
                   for i in range(0, len(indicadores_seg), BATCH_SIZE_F2)]

        n_prods = len(productos)
        print(f"  [{seg_idx:02d}/{total_segmentos}] {segmento[:48]}")
        print(f"           {len(indicadores_seg)} indicadores | {n_prods} productos | {len(batches)} lotes")

        for j, batch in enumerate(batches):
            print(f"    ⏳ Lote {j+1}/{len(batches)} ({len(batch)} ind.)...", end=" ", flush=True)
            prompt = prompt_fase2(batch, productos_txt, segmento)
            # max_tokens escala con el número de indicadores del lote
            data = llamar_api(client, prompt, max_tokens=max(600, len(batch) * 60))

            if data:
                results.extend(data)
                print(f"✅ {len(data)} ok")
            else:
                for ind in batch:
                    results.append({"indicador": ind, "producto": "Sin clasificar"})
                print("❌ fallido → 'Sin clasificar'")

            time.sleep(PAUSA)

    df_f2 = pd.DataFrame(results)[["indicador", "producto"]]
    sc = (df_f2["producto"] == "Sin clasificar").sum()
    print(f"\n  Fase 2 completa — clasificados: {len(df_f2)-sc}/{len(df_f2)}")
    return df_f2


# ══════════════════════════════════════════════
#  ENRIQUECIMIENTO Y EXPORTACIÓN

def enriquecer(df_f1, df_f2, df_jer):
    """
    Une Fase 1 (segmento) + Fase 2 (producto) y hace join
    con la jerarquía para obtener familia y clase.
    """
    df = df_f1.merge(df_f2, on="indicador", how="left")

    # Join con jerarquía para traer familia y clase
    lkp = (
        df_jer
        .drop_duplicates(subset="nombre_producto")
        [["nombre_producto", "nombre_familia", "nombre_clase"]]
    )
    df = df.merge(lkp, left_on="producto", right_on="nombre_producto", how="left")
    df = df.drop(columns=["nombre_producto"])

    for col in ["nombre_familia", "nombre_clase"]:
        df[col] = df[col].fillna("Sin clasificar")

    df = df.rename(columns={
        "indicador":      "Indicador",
        "segmento":       "Segmento",
        "producto":       "Producto",
        "nombre_familia": "Familia",
        "nombre_clase":   "Clase",
    })

    return df[["Indicador", "Segmento", "Familia", "Clase", "Producto"]]


def guardar(df_final, path):
    from openpyxl.styles import Font, PatternFill, Alignment

    col_widths = {"A": 70, "B": 45, "C": 45, "D": 45, "E": 55}

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_final.to_excel(writer, index=False, sheet_name="Clasificacion")
        ws = writer.sheets["Clasificacion"]

        fill  = PatternFill("solid", start_color="1E3A5F", end_color="1E3A5F")
        font  = Font(bold=True, name="Arial", size=11, color="FFFFFF")
        align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for letra, ancho in col_widths.items():
            ws.column_dimensions[letra].width = ancho
            cell = ws[f"{letra}1"]
            cell.font      = font
            cell.fill      = fill
            cell.alignment = align

        ws.row_dimensions[1].height = 22

    sc = (df_final["Producto"] == "Sin clasificar").sum()
    print(f"\n✅ Guardado: {path}")
    print(f"   Total indicadores   : {len(df_final)}")
    print(f"   Clasificados        : {len(df_final) - sc}")
    print(f"   Sin clasificar      : {sc}")

# ══════════════════════════════════════════════
#  MAIN

def main():
    print("=" * 55)
    print("  Clasificador 2 Fases — Indicadores → SECOP II")
    print("=" * 55)

    for f in [ARCHIVO_INDICADORES, ARCHIVO_JERARQUIA]:
        if not Path(f).exists():
            print(f"\n❌  Archivo no encontrado: {f}")
            return

    indicadores, df_jer = cargar_datos()
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    # ── Fase 1: indicador → segmento ──
    df_f1 = fase1(client, indicadores, df_jer)

    # ── Fase 2: indicador → producto (catálogo acotado por segmento) ──
    df_f2 = fase2(client, df_f1, df_jer)

    # ── Enriquecimiento con jerarquía completa ──
    df_final = enriquecer(df_f1, df_f2, df_jer)

    # ── Exportar ──
    guardar(df_final, ARCHIVO_SALIDA)


if __name__ == "__main__":
    main()
