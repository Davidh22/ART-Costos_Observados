"""
SECOP-II PDF Downloader — Indicadores PDET
==========================================
Lee los 4 archivos Excel de indicadores (GA1–GA4), extrae las URLs de SECOP-II
y descarga los Anexos Técnicos y Estudios Previos de cada contrato.

Estructura de carpetas de salida:
  descargas_indicadores/
    GA1/
      CO1.NTC.XXXXXXX/
        Estudio_Previo_...pdf
        Anexo_Tecnico_...pdf
    GA2/  GA3/  GA4/
  descargas_indicadores_FINAL.zip   ← archivo comprimido al terminar

USO:
  pip install selenium webdriver-manager pandas openpyxl
  python secop_descargador_indicadores.py

NOTAS:
  - GA1 y GA2: URLs en columna "URL SECOP II"
  - GA3 y GA4: URLs embebidas en columna "Nota de Costeo"
  - Se deduplican URLs repetidas entre archivos antes de procesar
  - Se genera informe Excel + ZIP comprimido al finalizar
"""

import os, re, time, shutil, logging, unicodedata, zipfile
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager

# ─── CONFIGURACIÓN ────────────────────────────────────────────────────────────
# Ajusta estas rutas a donde tengas los archivos Excel
EXCEL_FILES = {
    "GA1": "Indicadores_Final_GA1.xlsx",
    "GA2": "Indicadores_Final_GA2.xlsx",
    "GA3": "Indicadores_Final_GA3.xlsx",
    "GA4": "Indicadores_Final_GA4.xlsx",
}

OUTPUT_DIR   = "descargas_indicadores"   # carpeta raíz de descargas
REPORT_FILE  = "informe_descargas_PDET.xlsx"
ZIP_FILE     = "descargas_indicadores_FINAL.zip"
PROFILE_DIR  = os.path.abspath("chrome_secop_profile")
LOG_FILE     = "secop_indicadores.log"

DOWNLOAD_WAIT = 20   # seg. esperando que aparezca el archivo en disco
PAGE_WAIT     = 25   # seg. máximos para cargar la página SECOP

# ─── PALABRAS CLAVE PARA CLASIFICAR DOCUMENTOS ────────────────────────────────
KEYWORDS_ESTUDIO_PREVIO = [
    "estudio previo", "estudios previo", "estudios previos",
    "estudio de mercado", "estudios de mercado",
    "analisis del sector", "análisis del sector",
    "analisis de mercado", "análisis de mercado",
    "ep ", " ep.", "_ep_", "-ep-", "e.p.", "e.p ",
]

KEYWORDS_ANEXO_TECNICO = [
    "anexo tecnico", "anexos tecnicos", "anexo técnico", "anexos técnicos",
    "especificacion tecnica", "especificaciones tecnicas",
    "especificación técnica", "especificaciones técnicas",
    "ficha tecnica", "ficha técnica",
    "at ", " at.", "_at_", "-at-", "a.t.", "a.t ",
    "termino de referencia", "terminos de referencia",
    "término de referencia", "términos de referencia", "tdr",
]

KEYWORDS_SECUNDARIO = [
    "pliego de condicion", "pliego definitivo", "invitacion publica",
    "invitación pública", "resolucion de apertura", "acto administrativo",
    "minuta", "contrato", "adendo", "adición", "adicion",
]
# ──────────────────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger(__name__)


# ─── UTILIDADES ───────────────────────────────────────────────────────────────
def normalize(text: str) -> str:
    nfkd = unicodedata.normalize("NFKD", str(text))
    return "".join(c for c in nfkd if not unicodedata.combining(c)).lower().strip()

def safe_name(name: str) -> str:
    name = unicodedata.normalize("NFKD", str(name))
    name = "".join(c for c in name if not unicodedata.combining(c))
    name = re.sub(r'[<>:"/\\|?*\n\r\t]', "_", name)
    return name.strip(". ") or "sin_nombre"

def extract_notice_uid(url: str) -> str | None:
    m = re.search(r'noticeUID=([^&\s]+)', str(url))
    return m.group(1).strip() if m else None

def classify_document(filename: str) -> tuple:
    """Retorna (prioridad, etiqueta). 1=EP, 2=AT, 3=Secundario, 9=Otro."""
    n = normalize(filename)
    for kw in KEYWORDS_ESTUDIO_PREVIO:
        if kw in n:
            return 1, "ESTUDIO PREVIO"
    for kw in KEYWORDS_ANEXO_TECNICO:
        if kw in n:
            return 2, "ANEXO TÉCNICO"
    for kw in KEYWORDS_SECUNDARIO:
        if kw in n:
            return 3, "SECUNDARIO"
    return 9, "OTRO"


# ─── LECTURA Y EXTRACCIÓN DE URLs DESDE LOS EXCEL ────────────────────────────
def load_all_contracts() -> pd.DataFrame:
    """
    Lee los 4 Excel y devuelve un DataFrame unificado con columnas:
      grupo, codigo, indicador, url, notice_uid
    Deduplica por notice_uid.
    """
    records = []

    for grupo, path in EXCEL_FILES.items():
        if not os.path.exists(path):
            log.warning(f"Archivo no encontrado: {path} — se omite {grupo}")
            continue

        df = pd.read_excel(path)
        log.info(f"[{grupo}] {len(df)} filas leídas de {path}")

        # ── GA1 y GA2: columna "URL SECOP II" ──
        if "URL SECOP II" in df.columns:
            for _, row in df.iterrows():
                url = str(row.get("URL SECOP II", "")).strip()
                if not url.startswith("http"):
                    continue
                uid = extract_notice_uid(url)
                if uid:
                    records.append({
                        "grupo":    grupo,
                        "codigo":   str(row.get("Código", "")).strip(),
                        "indicador": str(row.get("Indicador", "")).strip()[:80],
                        "url":      url,
                        "notice_uid": uid,
                    })

        # ── GA3 y GA4: URLs embebidas en "Nota de Costeo" ──
        elif "Nota de Costeo" in df.columns:
            url_pattern = re.compile(r'https?://community\.secop\.gov\.co[^\s\'")\]]+')
            for _, row in df.iterrows():
                nota = str(row.get("Nota de Costeo", ""))
                found_urls = url_pattern.findall(nota)
                for url in found_urls:
                    uid = extract_notice_uid(url)
                    if uid:
                        records.append({
                            "grupo":    grupo,
                            "codigo":   str(row.get("Código", "")).strip(),
                            "indicador": str(row.get("Indicador", "")).strip()[:80],
                            "url":      url,
                            "notice_uid": uid,
                        })
        else:
            log.warning(f"[{grupo}] No se encontró columna de URL reconocida")

    df_all = pd.DataFrame(records)
    log.info(f"Total registros antes de deduplicar: {len(df_all)}")

    # Deduplicar por notice_uid (conservar primera aparición)
    df_all = df_all.drop_duplicates(subset=["notice_uid"]).reset_index(drop=True)
    log.info(f"Total contratos únicos a procesar: {len(df_all)}")
    return df_all


# ─── DRIVER DE CHROME ─────────────────────────────────────────────────────────
def build_driver(download_dir: str) -> webdriver.Chrome:
    opt = Options()
    opt.add_argument(f"--user-data-dir={PROFILE_DIR}")
    opt.add_argument("--profile-directory=Secop")
    opt.add_argument("--no-sandbox")
    opt.add_argument("--disable-dev-shm-usage")
    opt.add_argument("--disable-blink-features=AutomationControlled")
    opt.add_argument("--start-maximized")
    opt.add_experimental_option("excludeSwitches", ["enable-automation"])
    opt.add_experimental_option("useAutomationExtension", False)
    opt.add_experimental_option("prefs", {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
        "profile.default_content_settings.popups": 0,
    })
    svc = Service(ChromeDriverManager().install())
    drv = webdriver.Chrome(service=svc, options=opt)
    drv.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"
    })
    return drv


# ─── ESPERAR ARCHIVO NUEVO EN DISCO ───────────────────────────────────────────
def wait_for_new_file(watch_dir: Path, before_set: set, timeout=DOWNLOAD_WAIT) -> Path | None:
    deadline = time.time() + timeout
    while time.time() < deadline:
        time.sleep(0.8)
        current = set(watch_dir.glob("*"))
        new = current - before_set
        completed = [f for f in new if not str(f).endswith(".crdownload")]
        if completed:
            f = max(completed, key=lambda x: x.stat().st_mtime)
            size_before = f.stat().st_size
            time.sleep(1.5)
            size_after = f.stat().st_size
            if size_after >= size_before:
                return f
    return None


# ─── OBTENER LISTA DE DOCUMENTOS DE LA PÁGINA ────────────────────────────────
def get_document_list(driver: webdriver.Chrome) -> list[dict]:
    """
    Detecta documentos usando la estructura real de SECOP-II:
      Nombres:  #tdColumnDocumentNameP2Gen_spnDocumentName_N
      Botones:  #lnkDownloadLinkP3Gen_N
    """
    docs = []
    idx = 0
    while True:
        name_id   = f"tdColumnDocumentNameP2Gen_spnDocumentName_{idx}"
        button_id = f"lnkDownloadLinkP3Gen_{idx}"
        try:
            name_el = driver.find_element(By.ID, name_id)
            btn_el  = driver.find_element(By.ID, button_id)
            filename = name_el.text.strip()
            if filename:
                pri, label = classify_document(filename)
                docs.append({
                    "filename": filename,
                    "priority": pri,
                    "label":    label,
                    "element":  btn_el,
                    "idx":      idx,
                })
        except NoSuchElementException:
            break
        idx += 1
    docs.sort(key=lambda d: d["priority"])
    return docs


# ─── PROCESAR UN CONTRATO ─────────────────────────────────────────────────────
def process_contract(driver, grupo: str, codigo: str, notice_uid: str,
                     url: str, base_out: Path, tmp_dir: Path) -> dict:
    """
    Navega a la URL, detecta documentos prioritarios y los descarga
    en base_out / grupo / notice_uid /
    """
    # Carpeta destino: descargas_indicadores/GA1/CO1.NTC.XXXXXXX/
    dest_dir = base_out / grupo / safe_name(notice_uid)
    dest_dir.mkdir(parents=True, exist_ok=True)

    result = {
        "grupo": grupo,
        "codigo": codigo,
        "notice_uid": notice_uid,
        "url": url,
        "estado": "pendiente",
        "total_docs_pagina": 0,
        "docs_descargados": 0,
        "tiene_estudio_previo": "No",
        "tiene_anexo_tecnico": "No",
        "archivos_descargados": "",
        "error": "",
    }

    try:
        driver.get(url)
        try:
            WebDriverWait(driver, PAGE_WAIT).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
        except TimeoutException:
            pass

        time.sleep(4)
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(1)

        docs = get_document_list(driver)
        result["total_docs_pagina"] = len(docs)

        if not docs:
            result["estado"] = "sin_documentos"
            log.warning(f"  Sin documentos: {notice_uid}")
            return result

        # Filtrar: descargar EP y AT; si no hay ninguno, descargar Secundarios
        prioritarios = [d for d in docs if d["priority"] in (1, 2)]
        a_descargar  = prioritarios if prioritarios else [d for d in docs if d["priority"] == 3]
        # Si aún no hay nada, descargar todo lo disponible
        if not a_descargar:
            a_descargar = docs[:3]  # máximo 3 documentos "OTRO"

        downloaded_files = []

        for doc in a_descargar:
            before_set = set(tmp_dir.glob("*"))
            try:
                driver.execute_script("arguments[0].click();", doc["element"])
            except Exception:
                try:
                    doc["element"].click()
                except Exception as e:
                    log.warning(f"  No se pudo hacer clic en {doc['filename']}: {e}")
                    continue

            new_file = wait_for_new_file(tmp_dir, before_set)

            if new_file:
                # Nombre final: TIPO_DOCUMENTO_nombre_original.pdf
                tipo_prefix = doc["label"].replace(" ", "_")
                dest_name = f"{tipo_prefix}_{safe_name(doc['filename'])}"
                if not dest_name.lower().endswith(".pdf"):
                    dest_name += ".pdf"
                dest_path = dest_dir / dest_name

                shutil.move(str(new_file), str(dest_path))
                downloaded_files.append(dest_name)
                log.info(f"  ✓ [{doc['label']}] {dest_name}")

                if doc["priority"] == 1:
                    result["tiene_estudio_previo"] = "Sí"
                elif doc["priority"] == 2:
                    result["tiene_anexo_tecnico"] = "Sí"
            else:
                log.warning(f"  ✗ Timeout esperando: {doc['filename']}")

            time.sleep(1.5)

        result["docs_descargados"]   = len(downloaded_files)
        result["archivos_descargados"] = " | ".join(downloaded_files)
        result["estado"] = "exitoso" if downloaded_files else "descarga_fallida"

    except Exception as e:
        result["estado"] = "error"
        result["error"]  = str(e)[:200]
        log.error(f"  Error procesando {notice_uid}: {e}")

    return result


# ─── GENERAR INFORME EXCEL ────────────────────────────────────────────────────
def generate_report(results: list, path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Descargas"

    brd = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    headers = [
        "Grupo", "Código", "Notice UID", "Estado",
        "Docs en Página", "Docs Descargados",
        "Tiene Estudio Previo", "Tiene Anexo Técnico",
        "Archivos Descargados", "URL", "Error"
    ]
    hdr_fill = PatternFill("solid", start_color="1F4E79")
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(1, ci, h)
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.fill      = hdr_fill
        cell.border    = brd
        cell.alignment = Alignment(horizontal="center", vertical="center")

    status_color = {
        "exitoso":         "C6EFCE",
        "sin_documentos":  "FFEB9C",
        "descarga_fallida":"FFC7CE",
        "error":           "FFC7CE",
        "sin_url":         "D9D9D9",
        "pendiente":       "F2F2F2",
    }

    for ri, r in enumerate(results, 2):
        vals = [
            r["grupo"], r["codigo"], r["notice_uid"], r["estado"],
            r["total_docs_pagina"], r["docs_descargados"],
            r["tiene_estudio_previo"], r["tiene_anexo_tecnico"],
            r["archivos_descargados"], r["url"], r["error"]
        ]
        sfill = PatternFill("solid", start_color=status_color.get(r["estado"], "F2F2F2"))
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(ri, ci, v)
            cell.border    = brd
            cell.font      = Font(size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=(ci in [9, 10, 11]))
            if ci == 4:
                cell.fill = sfill

    col_widths = [8, 14, 22, 18, 14, 15, 18, 16, 55, 60, 40]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # ── Hoja resumen por grupo ──
    ws2 = wb.create_sheet("Resumen por Grupo")
    df  = pd.DataFrame(results)
    ws2["A1"] = "Resumen por Grupo de Análisis"
    ws2["A1"].font = Font(bold=True, size=12, color="1F4E79")
    ws2.merge_cells("A1:G1")

    h2 = ["Grupo", "Total", "Exitosos", "Sin Docs", "Falló Descarga", "Con EP", "Con AT"]
    for ci, h in enumerate(h2, 1):
        cell = ws2.cell(2, ci, h)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", start_color="1F4E79")
        cell.border = brd

    for ri, grp in enumerate(sorted(df["grupo"].unique()), 3):
        sub = df[df["grupo"] == grp]
        vals = [
            grp,
            len(sub),
            (sub["estado"] == "exitoso").sum(),
            (sub["estado"] == "sin_documentos").sum(),
            (sub["estado"] == "descarga_fallida").sum(),
            (sub["tiene_estudio_previo"] == "Sí").sum(),
            (sub["tiene_anexo_tecnico"]  == "Sí").sum(),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(ri, ci, v)
            cell.border = brd
            cell.font   = Font(size=10)

    # ── Hoja resumen general ──
    ws3 = wb.create_sheet("Resumen General")
    total = len(df)
    rows3 = [
        ("RESUMEN GENERAL", ""),
        ("Total contratos únicos procesados", total),
        ("Exitosos (≥1 descarga)",            (df["estado"] == "exitoso").sum()),
        ("Sin documentos en página",          (df["estado"] == "sin_documentos").sum()),
        ("Descarga fallida",                  (df["estado"] == "descarga_fallida").sum()),
        ("Error de navegación",               (df["estado"] == "error").sum()),
        ("", ""),
        ("DOCUMENTOS PRIORITARIOS", ""),
        ("Con Estudio Previo descargado",     (df["tiene_estudio_previo"] == "Sí").sum()),
        ("Con Anexo Técnico descargado",      (df["tiene_anexo_tecnico"]  == "Sí").sum()),
        ("Con al menos 1 prioritario",
         ((df["tiene_estudio_previo"] == "Sí") | (df["tiene_anexo_tecnico"] == "Sí")).sum()),
        ("Sin ningún doc prioritario",
         ((df["tiene_estudio_previo"] == "No") & (df["tiene_anexo_tecnico"] == "No")).sum()),
        ("", ""),
        ("Fecha del informe", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    title_fill = PatternFill("solid", start_color="1F4E79")
    for i, (label, val) in enumerate(rows3, 2):
        c1 = ws3.cell(i, 1, label)
        c2 = ws3.cell(i, 2, val)
        if label in ("RESUMEN GENERAL", "DOCUMENTOS PRIORITARIOS"):
            c1.font = Font(bold=True, color="FFFFFF", size=10)
            c1.fill = title_fill
            ws3.merge_cells(f"A{i}:B{i}")
        else:
            c1.font = Font(size=10, bold=bool(val != ""))
            c2.font = Font(size=10)
    ws3.column_dimensions["A"].width = 45
    ws3.column_dimensions["B"].width = 20

    wb.save(path)
    log.info(f"Informe guardado: {path}")


# ─── COMPRIMIR CARPETA DE DESCARGAS ──────────────────────────────────────────
def compress_downloads(source_dir: Path, zip_path: str):
    """Crea un ZIP de toda la carpeta de descargas (excluyendo _tmp_downloads)."""
    log.info(f"Comprimiendo {source_dir} → {zip_path} ...")
    total_files = 0
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED, compresslevel=6) as zf:
        for fpath in source_dir.rglob("*"):
            if "_tmp_downloads" in str(fpath):
                continue
            if fpath.is_file():
                arcname = fpath.relative_to(source_dir.parent)
                zf.write(fpath, arcname)
                total_files += 1
    size_mb = Path(zip_path).stat().st_size / 1_048_576
    log.info(f"ZIP creado: {zip_path} ({total_files} archivos, {size_mb:.1f} MB)")
    return total_files, size_mb


# ─── DIAGNÓSTICO INICIAL ──────────────────────────────────────────────────────
def run_diagnostic(driver, url: str, uid: str):
    print(f"\n{'='*60}")
    print(f"DIAGNÓSTICO — {uid}")
    print(f"{'='*60}")
    driver.get(url)
    try:
        WebDriverWait(driver, PAGE_WAIT).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
    except TimeoutException:
        pass
    time.sleep(5)
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(2)
    driver.execute_script("window.scrollTo(0, 0);")
    time.sleep(1)

    docs = get_document_list(driver)
    if docs:
        print(f"\n✅ {len(docs)} documento(s) detectado(s):\n")
        for d in docs:
            print(f"  [P{d['priority']}] [{d['label']:18s}] {d['filename']}")
        print("\nDetección OK. Presiona ENTER para procesar todos los contratos.")
    else:
        print("\n⚠️  No se detectaron documentos.")
        print("Causas posibles:")
        print("  1. CAPTCHA no resuelto")
        print("  2. Página no terminó de cargar")
        print("  3. Estructura de SECOP cambió")
        print("\nRevisa Chrome y presiona ENTER para continuar de todas formas.")
    input("\n>>> ENTER para continuar <<<\n")


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    log.info("=" * 60)
    log.info("SECOP-II Downloader — Indicadores PDET")
    log.info("=" * 60)

    # 1. Leer y unificar contratos
    df_contracts = load_all_contracts()
    if df_contracts.empty:
        log.error("No se encontraron URLs válidas en ninguno de los archivos Excel.")
        return

    print(f"\n{'='*60}")
    print(f"Contratos únicos a descargar: {len(df_contracts)}")
    print(f"Por grupo:")
    for g, cnt in df_contracts.groupby("grupo").size().items():
        print(f"  {g}: {cnt}")
    print(f"{'='*60}\n")

    # 2. Preparar carpetas
    tmp_dir  = Path(OUTPUT_DIR) / "_tmp_downloads"
    tmp_dir.mkdir(parents=True, exist_ok=True)
    base_out = Path(OUTPUT_DIR)
    for grupo in df_contracts["grupo"].unique():
        (base_out / grupo).mkdir(parents=True, exist_ok=True)

    # 3. Iniciar Chrome
    driver = build_driver(str(tmp_dir.resolve()))

    print("=" * 60)
    print("INSTRUCCIONES INICIALES")
    print("=" * 60)
    print("1. Chrome se abrirá con la primera URL de SECOP-II")
    print("2. Si aparece CAPTCHA, resuélvelo manualmente")
    print("3. Espera a que cargue la lista de documentos del contrato")
    print("4. Vuelve aquí y presiona ENTER para iniciar la descarga masiva")
    print("=" * 60)

    first_row = df_contracts.iloc[0]
    run_diagnostic(driver, first_row["url"], first_row["notice_uid"])

    # 4. Procesar contratos
    results = []
    total   = len(df_contracts)

    for idx, row in df_contracts.iterrows():
        num = idx + 1
        log.info(f"\n[{num}/{total}] {row['grupo']} — {row['notice_uid']}")

        res = process_contract(
            driver       = driver,
            grupo        = row["grupo"],
            codigo       = row["codigo"],
            notice_uid   = row["notice_uid"],
            url          = row["url"],
            base_out     = base_out,
            tmp_dir      = tmp_dir,
        )
        results.append(res)

        # Informe parcial cada 30 contratos
        if num % 30 == 0:
            parcial = REPORT_FILE.replace(".xlsx", f"_parcial_{num}.xlsx")
            generate_report(results, parcial)
            log.info(f"Informe parcial: {parcial}")

        time.sleep(2)

    # 5. Informe final
    driver.quit()
    generate_report(results, REPORT_FILE)

    # 6. Comprimir
    n_files, size_mb = compress_downloads(base_out, ZIP_FILE)

    # 7. Limpiar tmp
    shutil.rmtree(str(tmp_dir), ignore_errors=True)

    # 8. Resumen en consola
    df_res = pd.DataFrame(results)
    ok  = (df_res["estado"] == "exitoso").sum()
    nd  = (df_res["estado"] == "sin_documentos").sum()
    ep  = (df_res["tiene_estudio_previo"] == "Sí").sum()
    at  = (df_res["tiene_anexo_tecnico"]  == "Sí").sum()

    print(f"\n{'='*60}")
    print("RESUMEN FINAL")
    print(f"{'='*60}")
    print(f"Total contratos procesados:  {total}")
    print(f"Exitosos:                    {ok}")
    print(f"Sin documentos en página:    {nd}")
    print(f"Con Estudio Previo:          {ep}")
    print(f"Con Anexo Técnico:           {at}")
    print(f"\nInforme Excel: {REPORT_FILE}")
    print(f"PDFs en:       {OUTPUT_DIR}/")
    print(f"ZIP final:     {ZIP_FILE} ({n_files} archivos, {size_mb:.1f} MB)")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
